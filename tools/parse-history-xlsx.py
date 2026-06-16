#!/usr/bin/env python3
"""
Parse the legacy Zezet workbook (Camion_con_Jorge.xlsx) into clean JSON for the
TypeScript importer (prisma/import-history.ts).

Sources (verified in transcripts/_EXCEL_MIGRATION_RECON.md):
  - Inventarios  -> trucks (code, plate, year, sizeFt, purchaseDate, price, odometer)
  - Tournos      -> the ONLY per-trip source: client + charge + driver + pay +
                    helper + pay, per truck per date, up to 3 trips/truck/day,
                    plus per-truck-per-day fuel (Combustible). 2026-01-26..2026-06-08.

The output JSON carries real client financial data; it is gitignored and never
committed. Usage:
    python3 tools/parse-history-xlsx.py <workbook.xlsx> <out.json>
"""
import json
import re
import sys
from datetime import datetime

import openpyxl


def main() -> None:
    if len(sys.argv) != 3:
        sys.exit("usage: parse-history-xlsx.py <workbook.xlsx> <out.json>")
    xlsx, out = sys.argv[1], sys.argv[2]
    wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)

    trucks, truck_by_num = parse_trucks(wb["Inventarios"])
    clients, workers, logs = parse_tournos(wb["Tournos"], truck_by_num)

    # Any truck referenced in Tournos but missing from Inventarios -> add a stub
    # so the importer can satisfy the foreign key.
    have = {t["code"] for t in trucks}
    for l in logs:
        if l["truckCode"] not in have:
            trucks.append({"code": l["truckCode"], "plate": None, "year": None,
                           "sizeFt": None, "purchaseDate": None,
                           "purchasePrice": None, "odometerStart": None})
            have.add(l["truckCode"])

    # Per-month trip counts, for cross-checking against the Analisis sheet.
    per_month: dict[str, int] = {}
    trip_total = 0
    for l in logs:
        ym = l["date"][:7]
        n = len(l["trips"])
        per_month[ym] = per_month.get(ym, 0) + n
        trip_total += n
    dates = sorted(l["date"] for l in logs)

    payload = {
        "meta": {
            "trucks": len(trucks),
            "clients": len(clients),
            "workers": len(workers),
            "logs": len(logs),
            "trips": trip_total,
            "dateMin": dates[0] if dates else None,
            "dateMax": dates[-1] if dates else None,
            "perMonth": dict(sorted(per_month.items())),
        },
        "trucks": trucks,
        "clients": clients,
        "workers": workers,
        "logs": logs,
    }
    with open(out, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=1)
    wb.close()
    print(json.dumps(payload["meta"], ensure_ascii=False, indent=2))


def _num_str(v) -> str | None:
    """Two-decimal string for a numeric cell, else None."""
    if isinstance(v, (int, float)):
        return f"{float(v):.2f}"
    return None


def parse_trucks(ws):
    """Inventarios: header on row 3 (1-based), data rows 4..45."""
    rows = list(ws.iter_rows(values_only=True))
    trucks = []
    by_num: dict[str, str] = {}
    for r in rows[3:]:
        if not r:
            continue
        num = r[0]
        if num is None:
            continue
        numkey = str(int(num)) if isinstance(num, float) and num.is_integer() else str(num).strip()
        if numkey in ("", "None"):
            continue
        purchase, price, odo_compra, _odo_hoy, year, pies, placa = (
            r[1], r[2], r[3], r[4], r[5], r[6], r[7],
        )
        # Skip empty placeholder rows (number only, e.g. 36/37/38).
        if all(v is None for v in (purchase, price, odo_compra, year, pies, placa)):
            continue
        is_furgon = numkey.upper().startswith("F") or (
            isinstance(pies, str) and "FURGON" in pies.upper()
        )
        code = f"Furgón {numkey}" if is_furgon else f"Camión {numkey}"
        plate = None
        if placa is not None:
            plate = str(int(placa)) if isinstance(placa, float) and placa.is_integer() else str(placa).strip()
        trucks.append({
            "code": code,
            "plate": plate,
            "year": int(year) if isinstance(year, (int, float)) else None,
            "sizeFt": int(pies) if isinstance(pies, (int, float)) else None,
            "purchaseDate": purchase.date().isoformat() if isinstance(purchase, datetime) else None,
            "purchasePrice": _num_str(price),
            "odometerStart": int(odo_compra) if isinstance(odo_compra, (int, float)) else None,
        })
        by_num[numkey] = code
    return trucks, by_num


# Tournos row-type labels we care about, per recon §4.
WORKER_RE = re.compile(r"^\s*(\d+)\.\s*(.+?)\s*$")
TRUCK_RE = re.compile(r"Camion#(\w+)", re.I)


def _parse_worker(v):
    """'14. Jesus Ureta' -> {key, name}. Numeric/blank/0 -> None (no worker)."""
    if v is None or isinstance(v, (int, float)):
        return None
    s = str(v).strip()
    if s in ("", "0"):
        return None
    m = WORKER_RE.match(s)
    if m:
        return {"key": s, "name": m.group(2).strip()}
    return {"key": s, "name": s}


def parse_tournos(ws, truck_by_num):
    rows = list(ws.iter_rows(values_only=True))
    header = rows[0]

    # Truck columns: header label col -> truck number; value col = label col + 1.
    truck_cols: dict[int, str] = {}
    for ci, val in enumerate(header):
        if isinstance(val, str):
            m = TRUCK_RE.search(val)
            if m:
                truck_cols[ci] = m.group(1)

    def cell(row, i):
        return row[i] if row is not None and i < len(row) else None

    trip_labels = {"Tournos 1", "Conductor 1", "Ayudante 1",
                   "Tournos 2", "Conductor 2", "Ayudante 2",
                   "Tournos 3", "Salario conductor 3", "Salario ayudante 3",
                   "Combustible"}

    # Group rows into per-date blocks. A block begins at "Tournos 1"; the date is
    # in col A of that row (descending order in the sheet).
    blocks = []
    cur = None
    cur_date = None
    for r in rows:
        a, b = cell(r, 0), cell(r, 1)
        if isinstance(a, datetime):
            cur_date = a.date()
        lbl = b.strip() if isinstance(b, str) else None
        if lbl == "Tournos 1":
            cur = {"date": cur_date, "rows": {}}
            blocks.append(cur)
        if cur is not None and lbl in trip_labels:
            cur["rows"].setdefault(lbl, r)

    clients: dict[str, dict] = {}
    workers: dict[str, str] = {}
    logs = []

    slot_rows = {
        1: ("Tournos 1", "Conductor 1", "Ayudante 1"),
        2: ("Tournos 2", "Conductor 2", "Ayudante 2"),
        3: ("Tournos 3", "Salario conductor 3", "Salario ayudante 3"),
    }

    for blk in blocks:
        d = blk["date"]
        if d is None:
            continue
        rws = blk["rows"]
        for label_col, numkey in truck_cols.items():
            if numkey == "0":  # Camion#0 is not a real truck
                continue
            val_col = label_col + 1
            # Fuel (Combustible) is recorded in the truck's LABEL column, not the
            # value column where trip charges/pays go (verified: 2098 cells at the
            # label col, 0 at the value col).
            fuel = None
            if "Combustible" in rws:
                fuel = _num_str(cell(rws["Combustible"], label_col))
            trips = []
            for slot, (cl_lbl, dr_lbl, hp_lbl) in slot_rows.items():
                crow = rws.get(cl_lbl)
                if crow is None:
                    continue
                client = cell(crow, label_col)
                if client is None or isinstance(client, (int, float)):
                    continue
                cname = str(client).strip()
                if cname in ("", "0"):
                    continue
                drow, hrow = rws.get(dr_lbl), rws.get(hp_lbl)
                driver = _parse_worker(cell(drow, label_col)) if drow else None
                if driver is None:  # Trip requires a driver (FK not null)
                    continue
                helper = _parse_worker(cell(hrow, label_col)) if hrow else None
                clients.setdefault(cname, {"name": cname, "code": _client_code(cname, clients)})
                workers[driver["key"]] = driver["name"]
                if helper:
                    workers[helper["key"]] = helper["name"]
                trips.append({
                    "seq": slot,
                    "clientName": cname,
                    "charge": _num_str(cell(crow, val_col)) or "0.00",
                    "driverKey": driver["key"],
                    "driverPay": (_num_str(cell(drow, val_col)) if drow else None) or "0.00",
                    "helperKey": helper["key"] if helper else None,
                    "helperPay": (_num_str(cell(hrow, val_col)) if hrow else None) or "0.00",
                })
            if trips or fuel:
                logs.append({
                    "truckCode": truck_by_num.get(numkey, f"Camión {numkey}"),
                    "date": d.isoformat(),
                    "fuelCost": fuel,
                    "trips": trips,
                })

    return (
        list(clients.values()),
        [{"key": k, "fullName": v} for k, v in workers.items()],
        logs,
    )


def _client_code(name: str, existing: dict) -> str:
    base = re.sub(r"[^A-Za-z0-9]", "", name).upper()[:16] or "CLIENT"
    code = base
    n = 2
    used = {c["code"] for c in existing.values()}
    while code in used:
        code = f"{base}{n}"
        n += 1
    return code


if __name__ == "__main__":
    main()
